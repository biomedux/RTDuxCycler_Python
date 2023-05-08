import os
import time
import datetime

import openpyxl
import cv2

from pcr.hid.hid_controller import serial
from pcr.logger import log_start_expt, log_save_img, log_save_test_img

class PCRFileManager():
    ROOT_PATH = f"C:/mPCR"
    DATA_PATH = f"{ROOT_PATH}/data"

    def __init__(self):
        self.expts = []
        self.current_expt = None
        self.expt_data_path = None
        self.expt_img_path = None
        
        self.xlsx = None

    @staticmethod
    def gen_expt_name():
        return f"{serial}_{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}"
    
    def __set_values(self, expt_name, fluors):
        self.current_expt = expt_name
        self.expts.append(expt_name)

        self.expt_data_path = f"{self.DATA_PATH}/{expt_name}"
        self.expt_img_path = f"{self.expt_data_path}/img"

    @log_start_expt
    def start_task(self, expt_name, fluors):
        self.__set_values(expt_name, fluors)
        
        os.mkdir(self.expt_data_path)
        os.mkdir(self.expt_img_path)
        self.xlsx = PCRXlsx(expt_name, fluors)
    
    def end_task(self):
        self.current_expt = None
        self.expt_data_path = None
        self.expt_img_path = None
        
        self.xlsx = None

    @log_save_img
    def save_img(self, img, fluor, loop):
        cv2.imwrite(f"{self.expt_img_path}/{self.current_expt}_{fluor}_{loop}.png", img)

    @log_save_test_img
    def save_test_img(self, img, fluor):
        cv2.imwrite(f"{self.DATA_PATH}/img/{fluor}_test.png", img)


class PCRXlsx():
    HEADER = ["Cycle"] + list(range(1, 26))
    
    def __init__(self, expt_name, fluors):
        self.path = f"C:/mPCR/data/{expt_name}/{expt_name}.xlsx"
        self.fluors = fluors

        try:
            self.wb = openpyxl.load_workbook(f"C:/mPCR/data/{expt_name}/{expt_name}.xlsx")
            if not self.wb.sheetnames == fluors:
                self.create_xlsx()

        except FileNotFoundError as e:
            self.create_xlsx()
        
    def create_xlsx(self):
        self.wb = openpyxl.Workbook()
        self.wb.remove_sheet(self.wb.get_sheet_by_name("Sheet"))

        for ind, fluor in enumerate(self.fluors):
            self.wb.create_sheet(index=ind, title=fluor)
            self.wb[fluor].append(self.HEADER)
        self.wb.save(self.path)

    def record_values(self, fluor, values, cycle_no):
        self.wb[fluor].append([cycle_no] + values)
        self.wb.save(self.path)

    def close(self):
        self.wb.close()

    def __enter__(self):
        return self.wb

    def __exit__(self):
        self.close()

    def __del__(self):
        self.close()